"""
Export: Data Clean | Inventory Ready | Audit Trail (Method + Cluster ID) | Summary
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

THIN = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

METHOD_COLORS = {
    "CACHE":"E3F2FD","FUZZY_AUTO":"E8F5E9","FUZZY_SINGLE":"F1F8FF",
    "LLM":"FFF9C4","MANUAL_MERGE":"C8E6C9","AUTO_MERGED":"E8F5E9",
    "KEPT":"E3F2FD","UNREVIEWED":"FCE4EC","FALLBACK":"FFE0B2","UNPROCESSED":"F5F5F5",
}
METHOD_LABELS = {
    "CACHE":"💾 Cache","FUZZY_AUTO":"⚡ Fuzzy Auto","FUZZY_SINGLE":"✓ Fuzzy",
    "LLM":"🤖 AI","MANUAL_MERGE":"✋ Manual","AUTO_MERGED":"⚡ Auto",
    "KEPT":"🔒 Kept","UNREVIEWED":"⚠ Unreviewed","FALLBACK":"↩ Fallback",
}

def _hdr(cell,text,bg="1F4E79",fg="FFFFFF"):
    cell.value=text; cell.font=Font(bold=True,color=fg,name="Arial",size=10)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=THIN

def _cell(cell,value,number_format=None,bold=False,color=None,center=False,bg=None):
    cell.value=value; cell.font=Font(name="Arial",size=9,bold=bold,color=color or "000000")
    cell.border=THIN
    cell.alignment=Alignment(vertical="center",horizontal="center" if center else "left",wrap_text=False)
    if number_format: cell.number_format=number_format
    if bg: cell.fill=PatternFill("solid",start_color=bg)

def _build_inventory(final_rows,desc_col):
    clean_col=f"{desc_col} (Clean)"
    groups=defaultdict(list)
    for r in final_rows:
        groups[(str(r.get(clean_col) or r.get(desc_col,"")),str(r.get("Satuan","") or ""))].append(r)
    inv=[]
    for seq,((cd,sat),rows) in enumerate(sorted(groups.items(),key=lambda x:x[0][0]),1):
        tq=sum(float(r.get("Qty Order") or 0) for r in rows)
        th=sum(float(r.get("Total Harga (IDR)") or 0) for r in rows)
        hv=[float(r.get("Harga Satuan (IDR)") or 0) for r in rows if r.get("Harga Satuan (IDR)")]
        ah=sum(hv)/len(hv) if hv else 0
        kodes=sorted({str(r.get("Kode Material","")) for r in rows if r.get("Kode Material")})
        kats=sorted({str(r.get("Kategori","")) for r in rows if r.get("Kategori")})
        vendors=sorted({str(r.get("Vendor","")) for r in rows if r.get("Vendor")})
        pos=sorted({str(r.get("No PO","")) for r in rows if r.get("No PO")})
        sumber=sorted({str(r.get("Sumber RU","")) for r in rows if r.get("Sumber RU")})
        orig={str(r.get(desc_col,"")) for r in rows}
        methods=list({str(r.get("_method","")) for r in rows if r.get("_method")})
        inv.append({
            "No":seq,"Kode Material":" | ".join(kodes),"Deskripsi Material (Standar)":cd,
            "Satuan":sat,"Total Qty":tq,"Harga Satuan Rata-rata (IDR)":round(ah),
            "Total Harga (IDR)":round(th),"Kategori":" | ".join(kats),
            "Sumber RU":" | ".join(sumber),"Vendor":" | ".join(vendors),
            "Jumlah PO":len(pos),"No PO":" | ".join(pos),
            "Jumlah Baris Asal":len(rows),"Method":" | ".join(methods),
            "Status Merge":"✔ MERGED" if len(orig)>1 else "—",
            "_was_merged":len(orig)>1,
        })
    return inv

def export_results(final_rows,audit_entries,output_path,original_headers,metadata):
    try:
        wb=openpyxl.Workbook()
        desc_col=metadata.get("desc_col") or "Deskripsi Material"
        clean_col=f"{desc_col} (Clean)"

        # Sheet 1: Data Clean
        ws=wb.active; ws.title="Data Clean"
        skip={"_row_id","_normalized","_method","_cluster_id","_confidence"}
        oh=[h for h in original_headers if h not in skip]
        if desc_col in oh: oh.insert(oh.index(desc_col)+1,clean_col)
        else: oh.append(clean_col)
        oh+=["Method","Cluster ID"]
        for col,h in enumerate(oh,1):
            bg="2E7D32" if "(Clean)" in h else ("4A148C" if h in ("Method","Cluster ID") else "1F4E79")
            _hdr(ws.cell(1,col),h,bg=bg)
        for ri,rd in enumerate(final_rows,2):
            changed=str(rd.get(desc_col,""))!=str(rd.get(clean_col,""))
            method=str(rd.get("_method",""))
            alt="EAF4FB" if ri%2==0 else None
            for col,h in enumerate(oh,1):
                c=ws.cell(ri,col)
                if h=="Method": _cell(c,METHOD_LABELS.get(method,method),center=True,bg=METHOD_COLORS.get(method,alt))
                elif h=="Cluster ID": _cell(c,rd.get("_cluster_id",""),center=True,bg=alt)
                elif h in ("Harga Satuan (IDR)","Total Harga (IDR)"): _cell(c,rd.get(h),number_format="#,##0",bg=alt)
                elif h==clean_col and changed: _cell(c,rd.get(h),bold=True,color="1B5E20",bg="C8E6C9")
                else: _cell(c,rd.get(h),bg=alt)
        cw={"No":5,"No PO":22,"Tanggal PO":12,"Sumber RU":14,"Kode Material":14,
            desc_col:36,clean_col:40,"Satuan":8,"Qty Order":10,"Harga Satuan (IDR)":18,
            "Total Harga (IDR)":20,"Kategori":18,"Vendor":24,"Status PO":14,"Keterangan":28,
            "Method":16,"Cluster ID":10}
        for col,h in enumerate(oh,1): ws.column_dimensions[get_column_letter(col)].width=cw.get(h,16)
        ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(oh))}1"
        ws.row_dimensions[1].height=32

        # Sheet 2: Inventory Ready
        wi=wb.create_sheet("Inventory Ready")
        ih=["No","Kode Material","Deskripsi Material (Standar)","Satuan","Total Qty",
            "Harga Satuan Rata-rata (IDR)","Total Harga (IDR)","Kategori","Sumber RU",
            "Vendor","Jumlah PO","No PO","Jumlah Baris Asal","Method","Status Merge"]
        iw=[5,16,48,9,10,26,22,20,18,34,10,50,16,16,14]
        wi.merge_cells(f"A1:{get_column_letter(len(ih))}1")
        tc=wi.cell(1,1); tc.value="INVENTORY READY — MATERIAL UNIK TERSTANDARISASI"
        tc.font=Font(bold=True,name="Arial",size=12,color="FFFFFF")
        tc.fill=PatternFill("solid",start_color="1A237E")
        tc.alignment=Alignment(horizontal="center",vertical="center"); wi.row_dimensions[1].height=28
        for col,h in enumerate(ih,1): _hdr(wi.cell(2,col),h,bg="283593")
        wi.row_dimensions[2].height=30
        inventory=_build_inventory(final_rows,desc_col)
        for ri,item in enumerate(inventory,3):
            wm=item.pop("_was_merged",False)
            alt="EDE7F6" if wm else ("F8F9FA" if ri%2==0 else None)
            for col,h in enumerate(ih,1):
                c=wi.cell(ri,col); v=item.get(h)
                if h in ("Harga Satuan Rata-rata (IDR)","Total Harga (IDR)"): _cell(c,v,number_format="#,##0",center=True,bg=alt)
                elif h=="Total Qty": _cell(c,v,number_format="#,##0.##",center=True,bg=alt)
                elif h in ("No","Jumlah PO","Jumlah Baris Asal"): _cell(c,v,center=True,bg=alt)
                elif h=="Status Merge": _cell(c,v,bold=wm,color="1B5E20" if wm else "9E9E9E",center=True,bg=alt)
                else: _cell(c,v,bg=alt)
        for col,w in enumerate(iw,1): wi.column_dimensions[get_column_letter(col)].width=w
        lr=len(inventory)+3
        for col in range(1,len(ih)+1):
            c=wi.cell(lr,col); c.fill=PatternFill("solid",start_color="283593"); c.border=THIN
            c.font=Font(bold=True,name="Arial",size=9,color="FFFFFF")
            c.alignment=Alignment(horizontal="center",vertical="center")
        wi.cell(lr,1).value="TOTAL"
        qc=ih.index("Total Qty")+1; hc=ih.index("Total Harga (IDR)")+1; bc=ih.index("Jumlah Baris Asal")+1
        ql,hl,bl=get_column_letter(qc),get_column_letter(hc),get_column_letter(bc)
        wi.cell(lr,qc).value=f"=SUM({ql}3:{ql}{lr-1})"; wi.cell(lr,qc).number_format="#,##0.##"
        wi.cell(lr,hc).value=f"=SUM({hl}3:{hl}{lr-1})"; wi.cell(lr,hc).number_format="#,##0"
        wi.cell(lr,bc).value=f"=SUM({bl}3:{bl}{lr-1})"
        wi.freeze_panes="A3"; wi.auto_filter.ref=f"A2:{get_column_letter(len(ih))}2"

        # Sheet 3: Audit Trail
        wa=wb.create_sheet("Audit Trail")
        ah2=["Excel Row","No PO","Sumber RU","Deskripsi Original","Deskripsi Final (Clean)",
             "Aksi","Method","Cluster ID","Confidence","Similarity Score","Timestamp"]
        for col,h in enumerate(ah2,1): _hdr(wa.cell(1,col),h)
        wa.row_dimensions[1].height=30
        ac={"MERGED":"FFF9C4","MERGED (no change)":"E8F5E9","AUTO_MERGED":"E8F5E9",
            "KEPT":"E3F2FD","UNREVIEWED":"FCE4EC","FALLBACK":"FFE0B2"}
        for ri,e in enumerate(audit_entries,2):
            action=e.get("action",""); method=e.get("method","")
            bg=ac.get(action) or METHOD_COLORS.get(method)
            vals=[e.get("row_id"),e.get("no_po"),e.get("sumber_ru"),e.get("original"),
                  e.get("final"),action,METHOD_LABELS.get(method,method),
                  e.get("cluster_id"),e.get("confidence"),e.get("similarity_score"),e.get("timestamp")]
            for col,val in enumerate(vals,1):
                _cell(wa.cell(ri,col),val,bg=bg,center=(col in [1,8,9,10]))
        for col,w in enumerate([8,22,14,40,40,18,14,10,10,14,22],1):
            wa.column_dimensions[get_column_letter(col)].width=w
        wa.freeze_panes="A2"; wa.auto_filter.ref=f"A1:{get_column_letter(len(ah2))}1"
        wa.row_dimensions[1].height=30

        # Sheet 4: Summary
        ws3=wb.create_sheet("Summary")
        ws3['A1']="LAPORAN CLEANSING — LIST PO MATERIAL"
        ws3['A1'].font=Font(bold=True,name="Arial",size=14,color="1A237E")
        ws3.merge_cells("A1:D1")
        sr=[
            ("Tanggal Cleansing",datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("File Sumber",metadata.get("source_file","")),
            ("Kolom Duplikat",metadata.get("desc_col","Deskripsi Material")),
            ("Threshold",f"{metadata.get('threshold',75)}%"),
            ("",""),("— PIPELINE STATS —",""),
            ("Total Baris",metadata.get("total_rows",0)),
            ("Total Unik",metadata.get("total_unique",0)),
            ("Layer 2 Cache Hit",metadata.get("cache_hits",0)),
            ("Layer 3 Auto-merge",metadata.get("auto_merged",0)),
            ("Layer 4 AI Cluster",metadata.get("llm_clusters",0)),
            ("Layer 4 API Calls",metadata.get("llm_batches",0)),
            ("",""),("— HASIL —",""),
            ("Item Unik Inventory",len(inventory)),
            ("Total Cluster",metadata.get("total_clusters",0)),
            ("Duplikat",metadata.get("dup_clusters",0)),
            ("Manual Review",metadata.get("reviewed",0)),
            ("Manual Merge",metadata.get("merged",0)),
        ]
        for i,(lbl,val) in enumerate(sr,3):
            ws3.cell(i,1).value=lbl
            ws3.cell(i,1).font=Font(bold=lbl.startswith("—"),name="Arial",size=10,
                                    color="1A237E" if lbl.startswith("—") else "000000")
            ws3.cell(i,2).value=val; ws3.cell(i,2).font=Font(name="Arial",size=10)
            if lbl: ws3.cell(i,1).border=THIN; ws3.cell(i,2).border=THIN
        ws3.column_dimensions['A'].width=30; ws3.column_dimensions['B'].width=46
        leg=len(sr)+5
        ws3.cell(leg,1).value="Keterangan Method:"
        ws3.cell(leg,1).font=Font(bold=True,name="Arial",size=10,color="1A237E")
        for i,(m,c,d) in enumerate([
            ("💾 Cache","E3F2FD","Dari cache — tidak diproses ulang"),
            ("⚡ Fuzzy Auto","E8F5E9","Kemiripan ≥95% — auto-merge lokal"),
            ("🤖 AI","FFF9C4","Diputuskan oleh LLM (Claude)"),
            ("✋ Manual","C8E6C9","Diputuskan manual oleh user"),
            ("⚠ Unreviewed","FCE4EC","Belum direview saat export"),
            ("↩ Fallback","FFE0B2","API gagal — pakai nama original"),
        ],leg+1):
            c1=ws3.cell(i,1,m); c1.fill=PatternFill("solid",start_color=c)
            c1.font=Font(name="Arial",size=9,bold=True); c1.border=THIN
            ws3.cell(i,2).value=d; ws3.cell(i,2).font=Font(name="Arial",size=9)
            ws3.cell(i,2).border=THIN

        wb.save(output_path)
        return {"ok":True,
                "message":f"File berhasil disimpan:\n{output_path}\n\n"
                          f"• Data Clean — {len(final_rows)} baris\n"
                          f"• Inventory Ready — {len(inventory)} item unik\n"
                          f"• Audit Trail — {len(audit_entries)} entri\n• Summary",
                "inventory_count":len(inventory)}
    except Exception as e:
        import traceback
        return {"ok":False,"message":f"{str(e)}\n{traceback.format_exc()}"}
