"""
POCleaner — versi web, menggunakan pipeline_pg (PostgreSQL cache).
"""

import openpyxl
from datetime import datetime


class POCleaner:
    def __init__(self):
        from pipeline_pg import CleansingPipeline, normalize
        self._normalize = normalize
        self.raw_data = []
        self.clusters = []
        self.decisions = {}
        self.filepath = None
        self.header_row = []
        self.desc_col_index = None
        self.desc_col_name = None
        self.available_columns = []
        self._pipeline = CleansingPipeline()
        self._result_map = {}
        self._pipeline_stats = {}

    def load_excel(self, filepath):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else ""
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
            self.header_row = headers
            self.available_columns = [h for h in headers if h]
            self.raw_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row): continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                row_dict["_row_id"] = len(self.raw_data)
                self.raw_data.append(row_dict)
            self.filepath = filepath
            self.clusters = []; self.decisions = {}; self._result_map = {}
            self.desc_col_name = None; self.desc_col_index = None
            auto = None
            kw = ["deskripsi","description","material","nama","uraian","item","barang"]
            for h in self.available_columns:
                if any(k in h.lower() for k in kw): auto = h; break
            return {"ok": True, "message": f"{len(self.raw_data)} baris dimuat.",
                    "count": len(self.raw_data), "columns": self.available_columns,
                    "auto_detect": auto}
        except Exception as e:
            return {"ok": False, "message": str(e)}

    def set_desc_column(self, col_name):
        if col_name not in self.header_row:
            return {"ok": False, "message": f"Kolom '{col_name}' tidak ditemukan."}
        self.desc_col_name = col_name
        self.desc_col_index = self.header_row.index(col_name)
        for row in self.raw_data:
            val = row.get(col_name, "")
            row["_normalized"] = self._normalize(str(val) if val else "")
        self.clusters = []; self.decisions = {}; self._result_map = {}
        return {"ok": True}

    def run_pipeline(self, api_config, threshold=75, progress_cb=None, skip_llm=False):
        if not self.desc_col_name: return [], 0
        self._result_map = self._pipeline.run(
            self.raw_data, self.desc_col_name, api_config,
            progress_cb=progress_cb, skip_llm=skip_llm)
        self._pipeline_stats = self._pipeline.last_stats()
        self.clusters = self._convert_clusters(self._pipeline.last_clusters())
        return self.clusters, sum(1 for c in self.clusters if c["is_duplicate"])

    def _convert_clusters(self, raw_clusters):
        from collections import defaultdict
        desc_col = self.desc_col_name
        desc_to_rows = defaultdict(list)
        for r in self.raw_data:
            desc_to_rows[str(r.get(desc_col,"") or "")].append(r)

        result = []
        for cl in raw_clusters:
            members = []
            for orig in cl["orig_descs"]:
                members.extend(desc_to_rows.get(orig, []))

            canonical = cl["suggested_name"]
            method = "FUZZY_AUTO"; confidence = "high"
            reasoning = ""; warnings = []; is_same = True

            for nk in cl["norm_keys"]:
                rm = self._result_map.get(nk)
                if rm:
                    if rm.get("cleaned_text"): canonical = rm["cleaned_text"]
                    method     = rm.get("method", method)
                    confidence = rm.get("confidence", confidence)
                    reasoning  = rm.get("reasoning", "")
                    warnings   = rm.get("warnings", [])
                    is_same    = rm.get("is_same_item", True)
                    break

            ai_result = None
            if method in ("LLM", "CACHE") and canonical:
                ai_result = {"is_same_item": is_same, "confidence": confidence,
                             "canonical_name": canonical, "reasoning": reasoning,
                             "warnings": warnings}

            result.append({
                "id": hash(cl["cluster_id"]) % 100000,
                "cluster_id": cl["cluster_id"],
                "normalized_key": cl["norm_keys"][0] if cl["norm_keys"] else "",
                "items": members,
                "orig_descs": list(set(cl["orig_descs"])),
                "suggested_name": canonical,
                "avg_score": round(cl.get("min_score", 100), 1),
                "is_duplicate": cl["is_duplicate"],
                "is_auto": cl.get("is_auto", False),
                "method": method, "confidence": confidence,
                "ai_result": ai_result,
            })
        result.sort(key=lambda c: (not c["is_duplicate"], c["avg_score"]))
        return result

    def save_decision(self, cluster_id, action, canonical_name):
        self.decisions[cluster_id] = {
            "action": action, "canonical_name": canonical_name,
            "timestamp": datetime.now().isoformat()}

    def get_decision(self, cluster_id):
        return self.decisions.get(cluster_id)

    def apply_decisions(self):
        from pipeline_pg import normalize
        desc_col  = self.desc_col_name or "Deskripsi Material"
        clean_col = f"{desc_col} (Clean)"
        row_final = {}; audit_entries = []

        for cluster in self.clusters:
            decision = self.decisions.get(cluster["id"])
            cid = cluster.get("cluster_id", str(cluster["id"]))
            for item in cluster["items"]:
                rid  = item["_row_id"]
                orig = str(item.get(desc_col,"") or "")
                nk   = item.get("_normalized", normalize(orig))
                pr   = self._result_map.get(nk, {})
                pm   = pr.get("method","UNPROCESSED")
                pc   = pr.get("confidence","")
                if decision and decision["action"] == "merge":
                    final = decision["canonical_name"]; m = "MANUAL_MERGE"
                    al = "MERGED" if orig != final else "MERGED (no change)"
                elif cluster.get("is_auto") and not decision:
                    final = cluster["suggested_name"]; m = pm; al = "AUTO_MERGED"
                else:
                    final = orig; m = pm or "KEPT"
                    al = ("KEPT" if (decision and decision["action"]=="keep_all")
                          else ("AUTO_MERGED" if cluster.get("is_auto") else "UNREVIEWED"))
                row_final[rid] = (final, m, cid, pc)
                audit_entries.append({
                    "row_id": rid+2, "no_po": item.get("No PO",""),
                    "sumber_ru": item.get("Sumber RU",""),
                    "original": orig, "final": final, "action": al,
                    "method": m, "cluster_id": cid, "confidence": pc,
                    "similarity_score": cluster["avg_score"],
                    "timestamp": (decision or {}).get("timestamp",""),
                })

        final_rows = []
        for r in self.raw_data:
            rc = dict(r); rid = r["_row_id"]
            if rid in row_final:
                ft, m, cid, c = row_final[rid]
            else:
                ft = str(r.get(desc_col,"") or ""); m = "UNPROCESSED"; cid = ""; c = ""
            rc[clean_col] = ft; rc["_method"] = m
            rc["_cluster_id"] = cid; rc["_confidence"] = c
            final_rows.append(rc)
        return final_rows, audit_entries

    def stats(self):
        ps = self._pipeline_stats
        return {
            "total_rows": len(self.raw_data),
            "total_clusters": len(self.clusters),
            "dup_clusters": sum(1 for c in self.clusters if c["is_duplicate"]),
            "desc_col": self.desc_col_name,
            "cache_hits": ps.get("cache_hits",0),
            "auto_merged": ps.get("auto_merged",0),
            "llm_clusters": ps.get("llm_clusters",0),
            "llm_batches": ps.get("llm_batches",0),
        }

    def cache_stats(self):
        return self._pipeline.cache.stats()
